"""Turning an uploaded file into `Interaction` rows.

Shared by the ingest endpoint and by onboarding, which needs the same parse
before a BusinessProfile exists. Parsing and persistence are fast and happen in
the request; extraction is the slow part and runs behind a job.
"""

from __future__ import annotations

import hashlib
import uuid
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from app.ingestion.whatsapp_export import ParsedMessage, parse_text
from app.models.ingestion import Interaction
from app.services.storage import store_media

TEXT_SUFFIXES = {".txt"}
ZIP_SUFFIXES = {".zip"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}
# Voice notes. Gemini reads audio natively, so there is no transcription
# step to get wrong before the extractor sees it — which matters when the
# recording is Gujarati and Hindi with English quality codes in the middle.
AUDIO_SUFFIXES = {".ogg", ".oga", ".opus", ".m4a", ".mp3", ".wav", ".aac", ".webm"}

SUPPORTED = (TEXT_SUFFIXES | ZIP_SUFFIXES | EXCEL_SUFFIXES | IMAGE_SUFFIXES
             | AUDIO_SUFFIXES)

# A 90-day group export is a few thousand messages; well past that and the
# upload is more likely a mistake than a business.
MAX_MESSAGES = 50_000


class IntakeError(ValueError):
    """The upload cannot be read. Carries the status the API should return."""

    def __init__(self, status_code: int, detail: str):
        super().__init__(detail)
        self.status_code = status_code
        self.detail = detail


@dataclass
class Intake:
    kind: str  # whatsapp_export | excel | image
    interactions: list[Interaction]
    skipped: int
    media: int = 0


def dedupe_hash(
    tenant_id: uuid.UUID,
    thread: str | None,
    occurred_at,
    sender: str | None,
    body: str | None,
    media_file: str | None = None,
) -> str:
    """Identity of a message, so the same one never lands twice.

    Deliberately not a hash of the file. Uploading the identical export twice
    is the easy case; the one that actually happens is re-exporting the same
    chat a month later, where the file is different but most of the messages
    are ones we already hold. Hashing the message means the second upload
    contributes only what is new.

    The thread is deliberately NOT part of the key, even though it is passed
    in for readability at the call site. It is derived from the filename, and
    filenames are the least stable thing about an export: the same chat
    downloaded twice arrives as "chat.txt" and "chat (1).txt", and hashing
    that in would let the second copy through as new. Timestamp, sender and
    body already identify a message — two different chats carrying identical
    text from an identically-named sender in the same second is not a case
    worth protecting against at the cost of the one that happens weekly.
    """
    parts = [
        str(tenant_id),
        occurred_at.isoformat() if occurred_at else "",
        (sender or "").strip().lower(),
        (body or "").strip(),
        (media_file or "").strip().lower(),
    ]
    return hashlib.sha256("".join(parts).encode("utf-8")).hexdigest()


def _audio_mime(suffix: str) -> str:
    """What the browser actually recorded, so the model is told the truth.

    MediaRecorder gives ogg/opus on Firefox and webm/opus on Chrome, and
    guessing wrong makes Gemini reject the part rather than mis-read it.
    """
    return {
        ".ogg": "audio/ogg", ".oga": "audio/ogg", ".opus": "audio/ogg",
        ".webm": "audio/webm", ".m4a": "audio/mp4", ".mp3": "audio/mpeg",
        ".wav": "audio/wav", ".aac": "audio/aac",
    }.get(suffix, "audio/ogg")


def _phone_from_sender(sender: str | None) -> str | None:
    """Exports name the sender by contact name or by number, never both."""
    if not sender:
        return None
    digits = "".join(c for c in sender if c.isdigit())
    return digits[-10:] if len(digits) >= 10 else None


def _messages_from_zip(
    path: Path, tenant_id: uuid.UUID
) -> list[tuple[ParsedMessage, str | None]]:
    """WhatsApp's zip carries the transcript plus the photos and voice notes."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        txt = next((n for n in names if n.lower().endswith(".txt")), None)
        if txt is None:
            raise IntakeError(400, "Zip contains no chat transcript (.txt).")
        messages = list(parse_text(z.read(txt).decode("utf-8", errors="replace")))

        by_name = {Path(n).name: n for n in names}
        out: list[tuple[ParsedMessage, str | None]] = []
        for msg in messages:
            uri = None
            if msg.media_file and Path(msg.media_file).name in by_name:
                uri = store_media(
                    tenant_id, msg.media_file, z.read(by_name[Path(msg.media_file).name])
                )
            out.append((msg, uri))
        return out


def _rows_from_excel(path: Path) -> list[str]:
    """One row becomes one message.

    Crude on purpose: it reuses the extraction path that already works instead
    of inventing a second one. The real column-mapping importer is a connector
    (see BUILD_PROMPT section 6) and replaces this.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header = next(rows, None)
    if header is None:
        return []
    labels = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]

    bodies: list[str] = []
    for row in rows:
        pairs = [
            f"{labels[i]}: {value}"
            for i, value in enumerate(row)
            if i < len(labels) and value not in (None, "")
        ]
        if pairs:
            bodies.append("; ".join(pairs))
    workbook.close()
    return bodies


def interactions_from_upload(
    tenant_id: uuid.UUID, filename: str | None, path: Path, job_id: uuid.UUID
) -> Intake:
    """Parse an upload into unsaved `Interaction` rows tagged with the job."""
    suffix = Path(filename or "").suffix.lower()
    attributes: dict[str, Any] = {"job_id": str(job_id), "filename": filename}
    interactions: list[Interaction] = []
    skipped = 0

    if suffix in ZIP_SUFFIXES or suffix in TEXT_SUFFIXES:
        if suffix in ZIP_SUFFIXES:
            pairs = _messages_from_zip(path, tenant_id)
        else:
            pairs = [
                (m, None)
                for m in parse_text(path.read_text(encoding="utf-8", errors="replace"))
            ]

        if not pairs:
            raise IntakeError(400, "No messages found — is this a WhatsApp chat export?")
        if len(pairs) > MAX_MESSAGES:
            raise IntakeError(413, f"Export has {len(pairs)} messages; cap is {MAX_MESSAGES}.")

        thread = Path(filename or "chat").stem
        for msg, media_uri in pairs:
            if not (msg.body or media_uri):
                skipped += 1
                continue
            interactions.append(
                Interaction(
                    tenant_id=tenant_id,
                    channel="whatsapp_export",
                    sender=msg.sender,
                    sender_phone=_phone_from_sender(msg.sender),
                    occurred_at=msg.occurred_at,
                    body=msg.body,
                    media_uri=media_uri,
                    media_kind=msg.media_kind,
                    thread_key=thread,
                    dedupe_hash=dedupe_hash(
                        tenant_id, thread, msg.occurred_at, msg.sender,
                        msg.body, msg.media_file,
                    ),
                    attributes=attributes,
                )
            )
        media = sum(1 for _, uri in pairs if uri)
        return Intake("whatsapp_export", interactions, skipped, media)

    if suffix in EXCEL_SUFFIXES:
        bodies = _rows_from_excel(path)
        if not bodies:
            raise IntakeError(400, "Sheet has no data rows below the header.")
        for index, body in enumerate(bodies, start=2):  # row 1 is the header
            interactions.append(
                Interaction(
                    tenant_id=tenant_id,
                    channel="upload",
                    sender=filename,
                    body=body,
                    thread_key=f"{Path(filename or 'sheet').stem}#row{index}",
                    attributes=attributes,
                )
            )
        return Intake("excel", interactions, skipped)

    if suffix in AUDIO_SUFFIXES:
        uri = store_media(tenant_id, filename or "note.ogg", path.read_bytes())
        interactions.append(
            Interaction(
                tenant_id=tenant_id,
                channel="upload",
                sender=filename,
                body="",
                media_uri=uri,
                media_kind="audio",
                thread_key=Path(filename or "voice").stem,
                dedupe_hash=dedupe_hash(
                    tenant_id, None, None, filename, None, uri
                ),
                attributes={**attributes, "mime": _audio_mime(suffix)},
            )
        )
        return Intake("audio", interactions, 0, 1)

    if suffix in IMAGE_SUFFIXES:
        uri = store_media(tenant_id, filename or "photo.jpg", path.read_bytes())
        interactions.append(
            Interaction(
                tenant_id=tenant_id,
                channel="upload",
                sender=filename,
                body="",
                media_uri=uri,
                media_kind="image",
                attributes=attributes,
            )
        )
        return Intake("image", interactions, skipped)

    raise IntakeError(
        415,
        f"Unsupported file type '{suffix or filename}'. "
        "Send a WhatsApp export (.txt/.zip), an .xlsx sheet, or a photo.",
    )
