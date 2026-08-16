"""Establishing the party list before the first backfill.

This is what decides whether onboarding feels like magic or like homework. The
Resolver can only auto-commit a record it can attribute to somebody, so a
tenant with an empty `party` table sends its entire 90-day history to the
review queue — thousands of items, and the owner closes the app.

Sources, best first:

1. **Tally** — the accountant's own ledger master. Names, phones, GSTINs,
   credit terms and opening balances, already reconciled.
2. **Excel** — a customer list, in whatever column order the shop uses.
3. **The chat export itself** — who the owner actually talks to. Always
   available, which is the point: it is the floor, not the preference.

Imports are not extractions, so they do not go through `commit.py`'s
confidence gate — an accountant's ledger export is not a model guess. It is
still tenant-scoped and still deduplicated against what is already there.
"""

from __future__ import annotations

import re
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy import func, select

from app.models.finance import Invoice
from app.models.ingestion import Interaction
from app.models.party import Party
from app.services.matching import normalize_phone
from app.services.clock import business_today

# Tally groups every party under one of these. Anything else in the ledger
# master is a bank, a tax head, or an expense — not somebody who owes money.
CUSTOMER_GROUPS = ("sundry debtors", "trade receivables")
SUPPLIER_GROUPS = ("sundry creditors", "trade payables")

# Header names a shop might plausibly use, lowercased and stripped of spaces.
COLUMN_ALIASES = {
    "name": ("name", "party", "partyname", "customer", "customername", "ledger",
             "ledgername", "account", "accountname", "firm", "shopname"),
    "phone": ("phone", "mobile", "mobileno", "phoneno", "contact", "contactno",
              "whatsapp", "number"),
    "city": ("city", "place", "town", "location", "station"),
    "gstin": ("gstin", "gst", "gstno", "gstnumber", "tin"),
    "credit_days": ("creditdays", "credit", "days", "terms", "creditperiod"),
    "outstanding": ("outstanding", "balance", "openingbalance", "opening", "due",
                    "amount", "closingbalance", "baki"),
    "kind": ("kind", "type", "partytype", "category"),
}

# How far down to look for the header before giving up.
HEADER_SEARCH_ROWS = 15

# How much of each sheet the mapper is shown when the deterministic
# pass finds nothing. Kept small: this is a sample, not the data.
SAMPLE_ROWS = 12
SAMPLE_COLS = 20

OPENING_INVOICE_NO = "OPENING"


@dataclass
class PartySeed:
    name: str
    aliases: list[str] = field(default_factory=list)
    phone: str | None = None
    city: str | None = None
    gstin: str | None = None
    kind: str = "customer"
    credit_days: int | None = None
    outstanding: float | None = None


@dataclass
class ImportResult:
    source: str
    created: int = 0
    merged: int = 0
    skipped: int = 0
    opening_invoices: int = 0
    total_outstanding: float = 0.0

    def as_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "created": self.created,
            "merged": self.merged,
            "skipped": self.skipped,
            "opening_invoices": self.opening_invoices,
            "total_outstanding": round(self.total_outstanding, 2),
        }


# --- parsing --------------------------------------------------------------


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def _credit_days(value: Any) -> int | None:
    """Tally writes "45 Days"; a spreadsheet writes 45."""
    number = _number(value)
    return int(number) if number is not None and number >= 0 else None


def parse_tally_xml(content: bytes | str) -> list[PartySeed]:
    """Read parties out of a Tally ledger-master export.

    Tally exports debit balances as negative numbers, so a customer who owes
    money shows as -125000. Outstanding is stored as a magnitude and the group
    decides which direction it points.
    """
    if isinstance(content, bytes):
        # Tally writes ISO-8859-1 by default and declares it inconsistently.
        try:
            content = content.decode("utf-8")
        except UnicodeDecodeError:
            content = content.decode("iso-8859-1", errors="replace")

    content = re.sub(r"<\?xml[^>]*\?>", "", content, count=1).strip()
    try:
        root = ET.fromstring(content)
    except ET.ParseError as exc:
        raise ValueError(f"Not a readable Tally XML export: {exc}") from exc

    seeds: list[PartySeed] = []
    for ledger in root.iter("LEDGER"):
        name = _clean(ledger.get("NAME")) or _clean(ledger.findtext("NAME"))
        if not name:
            continue

        parent = (_clean(ledger.findtext("PARENT")) or "").lower()
        if any(group in parent for group in CUSTOMER_GROUPS):
            kind = "customer"
        elif any(group in parent for group in SUPPLIER_GROUPS):
            kind = "supplier"
        else:
            continue  # a bank, a tax head, an expense — not a party

        balance = _number(ledger.findtext("OPENINGBALANCE"))
        mailing = [_clean(n.text) for n in ledger.iter("MAILINGNAME") if _clean(n.text)]

        seeds.append(
            PartySeed(
                name=name,
                aliases=[a for a in mailing if a.lower() != name.lower()],
                phone=(
                    _clean(ledger.findtext("LEDGERMOBILE"))
                    or _clean(ledger.findtext("LEDGERPHONE"))
                    or _clean(ledger.findtext("LEDGERCONTACT"))
                ),
                city=_clean(ledger.findtext("LEDGERSTATENAME")) or _clean(
                    ledger.findtext(".//ADDRESS")
                ),
                gstin=(
                    _clean(ledger.findtext("PARTYGSTIN"))
                    or _clean(ledger.findtext("INCOMETAXNUMBER"))
                ),
                kind=kind,
                credit_days=_credit_days(ledger.findtext("BILLCREDITPERIOD")),
                outstanding=abs(balance) if balance else None,
            )
        )
    return seeds


def _column_map(header: tuple) -> dict[str, int]:
    """Match spreadsheet headers to fields, whatever order they came in."""
    found: dict[str, int] = {}
    for index, cell in enumerate(header):
        key = re.sub(r"[^a-z]", "", str(cell or "").lower())
        if not key:
            continue
        for target, aliases in COLUMN_ALIASES.items():
            if target not in found and key in aliases:
                found[target] = index
                break
    return found


def _sheet_sample(sheet, rows: int, cols: int) -> list[list]:
    """The first few rows of a sheet, for the model to look at."""
    out: list[list] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index >= rows:
            break
        out.append([("" if c is None else str(c))[:60] for c in row[:cols]])
    return out


def _seeds_from_grid(grid: list[tuple], columns: dict[str, int], start_row: int) -> list[PartySeed]:
    """Read parties out of already-loaded rows using a column mapping."""
    def cell(row, key):
        index = columns.get(key)
        return row[index] if index is not None and index < len(row) else None

    seeds: list[PartySeed] = []
    for row in grid[start_row:]:
        name = _clean(cell(row, "name"))
        if not name:
            continue
        kind = (_clean(cell(row, "kind")) or "customer").lower()
        seeds.append(
            PartySeed(
                name=name,
                phone=_clean(cell(row, "phone")),
                city=_clean(cell(row, "city")),
                gstin=_clean(cell(row, "gstin")),
                kind="supplier" if kind.startswith("suppl") else "customer",
                credit_days=_credit_days(cell(row, "credit_days")),
                outstanding=_number(cell(row, "outstanding")),
            )
        )
    return seeds


def parse_party_excel(path: Path, db=None, tenant_id: uuid.UUID | None = None) -> list[PartySeed]:
    """Read a customer list. Column order, casing and sheet order are the shop's business.

    Two passes, cheapest first. The deterministic one tries every sheet against
    the known heading spellings and costs nothing. Only when every sheet fails
    is the model asked where the party list is — because a book whose customer
    list is the third sheet under a title block is a normal book, not a broken
    one, and refusing it taught the owner that the product is fussy.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = list(workbook.worksheets)

        # --- pass one: score every sheet that could be a party list --------
        # First-match-wins is wrong here. An invoice sheet has a "Party"
        # column too, and picking it yields one row per *invoice* — the same
        # customer five times, and four sheets of real customers ignored.
        # So gather every candidate and rank them.
        samples: list[dict] = []
        candidates: list[tuple[float, list[PartySeed]]] = []

        for sheet in sheets:
            grid = list(sheet.iter_rows(values_only=True))
            samples.append({"name": sheet.title, "rows": [
                [("" if c is None else str(c))[:60] for c in row[:SAMPLE_COLS]]
                for row in grid[:SAMPLE_ROWS]
            ]})

            for index, row in enumerate(grid[:HEADER_SEARCH_ROWS]):
                columns = _column_map(row)
                if "name" not in columns:
                    continue
                seeds = _seeds_from_grid(grid, columns, index + 1)
                if not seeds:
                    continue

                score = 0.0
                # A sheet that says what it is, is usually telling the truth.
                title = re.sub(r"[^a-z]", "", sheet.title.lower())
                if any(w in title for w in
                       ("customer", "party", "ledger", "supplier", "vendor",
                        "master", "account", "client")):
                    score += 2.0
                if any(w in title for w in ("invoice", "bill", "sale", "purchase",
                                            "stock", "item", "payment", "voucher")):
                    score -= 2.0
                # One row per party: a party list has distinct names, a
                # transaction list repeats them.
                names = [seed.name.lower() for seed in seeds]
                score += 2.0 * (len(set(names)) / len(names))
                # Columns only a party list tends to carry.
                score += sum(0.5 for f in ("phone", "gstin", "credit_days", "city")
                             if f in columns)
                candidates.append((score, seeds))
                break

        candidates.sort(key=lambda c: c[0], reverse=True)
        # Confident only when there is a clear winner. A near-tie means two
        # sheets look equally like the party list, which is exactly when a
        # human eye — or the model — is worth the call.
        if candidates and (
            len(candidates) == 1 or candidates[0][0] - candidates[1][0] >= 1.0
        ) and candidates[0][0] >= 2.0:
            return candidates[0][1]

        # --- pass two: ask where it is -------------------------------------
        if db is not None and tenant_id is not None:
            from app.agents.sheet_mapper import SheetMapper

            decision = SheetMapper(db, tenant_id).execute({"sheets": samples})
            mapped = decision.output or {}
            if mapped.get("found"):
                target = next(
                    (sh for sh in sheets if sh.title == mapped.get("sheet")),
                    sheets[0] if sheets else None,
                )
                if target is not None:
                    grid = list(target.iter_rows(values_only=True))
                    header_row = mapped.get("header_row", -1)
                    seeds = _seeds_from_grid(
                        grid, mapped["columns"], max(0, header_row + 1)
                    )
                    if seeds:
                        return seeds

        # The model declined, failed, or was never available. Fall back to the
        # deterministic guess ONLY if it looked like a party list on its own
        # merits. A negative score means the only thing we could read was an
        # invoice sheet, and returning one row per invoice as the customer
        # list is worse than saying we could not read the file.
        if candidates and candidates[0][0] >= 1.0:
            return candidates[0][1]

        seen = sorted({
            str(c).strip()
            for sample in samples for row in sample["rows"] for c in row
            if c not in (None, "")
        })[:12]
        raise ValueError(
            "No list of customers or suppliers found in this workbook. One "
            "column needs to hold party names — headings like: "
            + ", ".join(COLUMN_ALIASES["name"])
            + (f". What was found instead: {', '.join(seen)}" if seen
               else ". The workbook looks empty.")
        )
    finally:
        workbook.close()


def parse_upload(
    filename: str | None, path: Path, db=None, tenant_id: uuid.UUID | None = None
) -> tuple[str, list[PartySeed]]:
    """db and tenant_id are optional so the parsers stay unit-testable without
    a session. Passing them enables the model fallback for oddly-shaped books;
    without them the deterministic pass is all there is."""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xml":
        return "tally", parse_tally_xml(path.read_bytes())
    if suffix in {".xlsx", ".xlsm"}:
        return "excel", parse_party_excel(path, db=db, tenant_id=tenant_id)
    raise ValueError(
        f"Unsupported party list '{suffix or filename}'. "
        "Send a Tally XML master export or an .xlsx customer list."
    )


# --- the fallback source: the chat itself --------------------------------

# Names an export uses for the owner's own side of the conversation.
_SELF_NAMES = {"you", "me", "self", "aap"}


def _tokens(text: str | None) -> set[str]:
    """Significant words in a business name, for owner matching.

    "Ravi Fabrics & Bros" and a sender called "Mahesh Shah (Ravi Fabrics)"
    are the same side of the conversation; comparing whole strings would miss
    that, and comparing every word would match on "textiles".
    """
    if not text:
        return set()
    stop = {"and", "bros", "brothers", "co", "company", "sons", "the",
            "textiles", "textile", "mills", "mill", "traders", "trading",
            "industries", "enterprise", "enterprises", "pvt", "ltd", "llp"}
    words = re.findall(r"[a-z0-9]+", text.lower())
    return {w for w in words if len(w) > 2 and w not in stop}


def is_owner(sender: str | None, phone: str | None, tenant) -> bool:
    """Whether this sender is the tenant themselves rather than a customer.

    A 1:1 export contains exactly two people, and one of them is the owner.
    Seeding both makes the owner a customer of themselves — measured on a real
    export, that put half of all records on the wrong party and would have
    shown the owner money they owe to themselves.
    """
    if not sender:
        return False
    if sender.strip().lower() in _SELF_NAMES:
        return True
    if tenant is None:
        return False

    if phone and normalize_phone(phone) and normalize_phone(phone) == normalize_phone(
        getattr(tenant, "owner_phone", None)
    ):
        return True

    theirs = _tokens(sender)
    if not theirs:
        return False

    # The names must match outright, or one must contain the other on the
    # strength of at least two words. A customer called "Ashok Bhai" contains
    # an owner called "Ashok", and dropping them from the party list is worse
    # than seeding one party too many: an unseeded customer sends every record
    # about them to review.
    for ours in (
        _tokens(getattr(tenant, "business_name", None)),
        _tokens(getattr(tenant, "owner_name", None)),
    ):
        if not ours:
            continue
        if theirs == ours:
            return True
        if (theirs <= ours or ours <= theirs) and min(len(theirs), len(ours)) >= 2:
            return True
    return False


def seeds_from_messages(
    db, tenant_id: uuid.UUID, owner_phone: str | None = None, tenant=None
) -> list[PartySeed]:
    """Every distinct counterparty the owner has actually talked to.

    A 1:1 export names the sender by their saved contact name; a group export
    names them by number. Both become a party, and the phone is what lets the
    Resolver attribute a message whose text never says who it is from.

    The owner's own side is excluded — see `is_owner`.
    """
    from app.models.tenant import Tenant

    if tenant is None:
        tenant = db.get(Tenant, tenant_id)
    rows = db.execute(
        select(
            Interaction.sender,
            Interaction.sender_phone,
            func.count().label("messages"),
        )
        .where(Interaction.tenant_id == tenant_id, Interaction.sender.isnot(None))
        .group_by(Interaction.sender, Interaction.sender_phone)
        .order_by(func.count().desc())
    ).all()

    owner_last10 = normalize_phone(owner_phone)
    merged: dict[str, PartySeed] = {}

    for sender, phone, _count in rows:
        name = _clean(sender)
        if not name or name.lower() in _SELF_NAMES:
            continue
        if owner_last10 and normalize_phone(phone) == owner_last10:
            continue  # the owner's own messages
        if is_owner(name, phone, tenant):
            continue

        key = name.lower()
        if key in merged:
            merged[key].phone = merged[key].phone or phone
        else:
            merged[key] = PartySeed(name=name, phone=phone, kind="customer")

    return list(merged.values())


# --- writing --------------------------------------------------------------


def _existing_match(db, tenant_id: uuid.UUID, seed: PartySeed) -> Party | None:
    """Name first, then phone — the two things that are never coincidences."""
    from app.services.matching import exact_alias_match, phone_match

    return (
        exact_alias_match(db, tenant_id, seed.name)
        or phone_match(db, tenant_id, seed.phone)
    )


def import_parties(
    db,
    tenant_id: uuid.UUID,
    seeds: list[PartySeed],
    source: str,
    *,
    opening_balances: bool = True,
) -> ImportResult:
    """Create or merge parties, and turn any outstanding into an open invoice."""
    result = ImportResult(source=source)
    today = business_today()

    for seed in seeds:
        if not seed.name:
            result.skipped += 1
            continue

        party = _existing_match(db, tenant_id, seed)
        if party is None:
            party = Party(
                tenant_id=tenant_id,
                name=seed.name,
                aliases=list(seed.aliases),
                phone=seed.phone,
                city=seed.city,
                gstin=seed.gstin,
                kind=seed.kind,
                credit_days=seed.credit_days or 0,
                source=source,
            )
            db.add(party)
            db.flush()
            result.created += 1
        else:
            # Never overwrite what is already known — an import is new
            # information, not a more authoritative version of old information.
            party.phone = party.phone or seed.phone
            party.city = party.city or seed.city
            party.gstin = party.gstin or seed.gstin
            if seed.credit_days and not party.credit_days:
                party.credit_days = seed.credit_days
            known = {a.lower() for a in (party.aliases or [])} | {party.name.lower()}
            extra = [a for a in [*seed.aliases, seed.name] if a.lower() not in known]
            if extra:
                party.aliases = [*(party.aliases or []), *extra]
            result.merged += 1

        if opening_balances and seed.outstanding and seed.outstanding > 0:
            already = db.execute(
                select(func.count())
                .select_from(Invoice)
                .where(
                    Invoice.tenant_id == tenant_id,
                    Invoice.party_id == party.id,
                    Invoice.invoice_no == OPENING_INVOICE_NO,
                )
            ).scalar_one()
            if not already:
                db.add(
                    Invoice(
                        tenant_id=tenant_id,
                        party_id=party.id,
                        invoice_no=OPENING_INVOICE_NO,
                        invoice_date=today,
                        # Due now, not overdue. A ledger-master export carries
                        # the balance but not the bills behind it, so its real
                        # age is unknown — and inventing an age would put
                        # numbers in the ageing buckets that the owner cannot
                        # reconcile against Tally. Bill-wise import is what
                        # fixes this properly.
                        due_date=today,
                        amount=seed.outstanding,
                        status="open",
                        source=source,
                    )
                )
                result.opening_invoices += 1
                result.total_outstanding += seed.outstanding

    db.flush()
    return result
