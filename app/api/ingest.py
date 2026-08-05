"""Ingestion routes.

The upload is the onboarding magic trick: the owner exports a chat, and a few
minutes later their last 90 days of orders and outstandings are on screen. So
this endpoint does the fast, safe half synchronously — parse and persist the
raw `Interaction` rows — and hands the slow half (a model call per message) to
a background job. The owner gets a job id back before their phone locks.
"""

from __future__ import annotations

import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, File, HTTPException, UploadFile

from app.api.deps import Profile, TenantDB, TenantId
from app.ingestion.whatsapp_export import ParsedMessage, parse_text
from app.models.ingestion import Interaction
from app.schemas.ingest import IngestAccepted, JobStatus
from app.services.backfill import job_status, run_backfill
from app.services.storage import store_media

router = APIRouter()

TEXT_SUFFIXES = {".txt"}
ZIP_SUFFIXES = {".zip"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}

# A 90-day group export is a few thousand messages; well past that and the
# upload is more likely a mistake than a business.
MAX_MESSAGES = 50_000


def _messages_from_zip(path: Path, tenant_id: uuid.UUID) -> list[tuple[ParsedMessage, str | None]]:
    """WhatsApp's zip carries the transcript plus the photos and voice notes."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        txt = next((n for n in names if n.lower().endswith(".txt")), None)
        if txt is None:
            raise HTTPException(400, "Zip contains no chat transcript (.txt).")
        messages = list(parse_text(z.read(txt).decode("utf-8", errors="replace")))

        by_name = {Path(n).name: n for n in names}
        out: list[tuple[ParsedMessage, str | None]] = []
        for msg in messages:
            uri = None
            if msg.media_file and Path(msg.media_file).name in by_name:
                uri = store_media(
                    tenant_id, msg.media_file, z.read(by_name[Path(msg.media_file).name])
                )
            out.append((msg, uri))
        return out


def _rows_from_excel(path: Path) -> list[str]:
    """One row becomes one message.

    Crude on purpose: it reuses the extraction path that already works instead
    of inventing a second one. The real column-mapping importer is a connector
    (see BUILD_PROMPT section 6) and replaces this.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header = next(rows, None)
    if header is None:
        return []
    labels = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)]

    bodies: list[str] = []
    for row in rows:
        pairs = [
            f"{labels[i]}: {value}"
            for i, value in enumerate(row)
            if i < len(labels) and value not in (None, "")
        ]
        if pairs:
            bodies.append("; ".join(pairs))
    workbook.close()
    return bodies


def _phone_from_sender(sender: str | None) -> str | None:
    """Exports name the sender by contact name or by number, never both."""
    if not sender:
        return None
    digits = "".join(c for c in sender if c.isdigit())
    return digits[-10:] if len(digits) >= 10 else None


@router.post("", response_model=IngestAccepted, status_code=202)
@router.post("/", response_model=IngestAccepted, status_code=202, include_in_schema=False)
def ingest_upload(
    tid: TenantId,
    db: TenantDB,
    profile: Profile,
    background: BackgroundTasks,
    file: UploadFile = File(...),
) -> IngestAccepted:
    """Accept a WhatsApp export (.txt/.zip), an Excel sheet, or a photo."""
    suffix = Path(file.filename or "").suffix.lower()
    job_id = uuid.uuid4()
    attributes: dict[str, Any] = {"job_id": str(job_id), "filename": file.filename}

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = Path(tmp.name)

    try:
        interactions: list[Interaction] = []
        skipped = 0

        if suffix in ZIP_SUFFIXES or suffix in TEXT_SUFFIXES:
            kind = "whatsapp_export"
            if suffix in ZIP_SUFFIXES:
                pairs = _messages_from_zip(tmp_path, tid)
            else:
                content = tmp_path.read_text(encoding="utf-8", errors="replace")
                pairs = [(m, None) for m in parse_text(content)]

            if not pairs:
                raise HTTPException(400, "No messages found — is this a WhatsApp chat export?")
            if len(pairs) > MAX_MESSAGES:
                raise HTTPException(
                    413, f"Export has {len(pairs)} messages; cap is {MAX_MESSAGES}."
                )

            thread = Path(file.filename or "chat").stem
            for msg, media_uri in pairs:
                if not (msg.body or media_uri):
                    skipped += 1
                    continue
                interactions.append(
                    Interaction(
                        tenant_id=tid,
                        channel="whatsapp_export",
                        sender=msg.sender,
                        sender_phone=_phone_from_sender(msg.sender),
                        occurred_at=msg.occurred_at,
                        body=msg.body,
                        media_uri=media_uri,
                        media_kind=msg.media_kind,
                        thread_key=thread,
                        attributes=attributes,
                    )
                )

        elif suffix in EXCEL_SUFFIXES:
            kind = "excel"
            bodies = _rows_from_excel(tmp_path)
            if not bodies:
                raise HTTPException(400, "Sheet has no data rows below the header.")
            for index, body in enumerate(bodies, start=2):  # row 1 is the header
                interactions.append(
                    Interaction(
                        tenant_id=tid,
                        channel="upload",
                        sender=file.filename,
                        body=body,
                        thread_key=f"{Path(file.filename or 'sheet').stem}#row{index}",
                        attributes=attributes,
                    )
                )

        elif suffix in IMAGE_SUFFIXES:
            kind = "image"
            uri = store_media(tid, file.filename or "photo.jpg", tmp_path.read_bytes())
            interactions.append(
                Interaction(
                    tenant_id=tid,
                    channel="upload",
                    sender=file.filename,
                    body="",
                    media_uri=uri,
                    media_kind="image",
                    attributes=attributes,
                )
            )

        else:
            raise HTTPException(
                415,
                f"Unsupported file type '{suffix or file.filename}'. "
                "Send a WhatsApp export (.txt/.zip), an .xlsx sheet, or a photo.",
            )
    finally:
        tmp_path.unlink(missing_ok=True)

    db.add_all(interactions)
    db.flush()

    # Commit before queueing: the background task opens its own session and
    # must not race the request's transaction.
    db.commit()
    background.add_task(run_backfill, tid, job_id)

    return IngestAccepted(
        job_id=job_id,
        interactions=len(interactions),
        skipped=skipped,
        kind=kind,
        detail=f"{len(interactions)} interactions stored; extraction running in the background.",
    )


@router.get("/jobs/{job_id}", response_model=JobStatus)
def get_job(job_id: uuid.UUID, tid: TenantId, db: TenantDB) -> JobStatus:
    status = job_status(db, tid, job_id)
    if status["total"] == 0:
        raise HTTPException(404, f"No ingestion job {job_id} for this tenant.")
    return JobStatus(**status)
